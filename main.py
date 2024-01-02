import openpyxl
import json

book = openpyxl.load_workbook("data.xlsx", read_only=True)

sheet = book.active

num = input('123: ')

status = "Number not found!"

d = {}

for item in range(1, sheet.max_row):
    if sheet[item][0].value == num:
        for item2 in range(0, sheet.max_column):
            if sheet[item][item2].value:
                # await bot.send_message(message.chat.id, f"{sheet[4][item2].value}: {sheet[item][item2].value}")
                print(sheet[3][item2].value, ': ', sheet[item][item2].value)
        status = "Found"

# key_words = sorted(d.items(), reverse=True)

d_str = json.dumps(d, ensure_ascii=False, indent=4)

print(d_str)

print(status if status == "Number not found!" else ' ')

# +998935260007
