import openpyxl
import os
import shutil
from zipfile import ZipFile
from utils.database import Database

db = Database("data/db.sqlite3")


def add_plus(number: str) -> str:
    """
        gets the user's phone number and checks if there is a plus at the beginning,
        if not then adds a plus and returns
    """
    if number.startswith("+"):
        return number
    return "+" + number


def get_salary(number, user_id, username, lastname):
    """
        accepts a phone number,
        opens the xlsx file and goes through the cell of phone numbers,
        if it finds a similar number, it parses its data and displays it in a beautiful form
    """
    # Создаем временную папку
    tmp_folder = 'tmp/convert_wrong_excel/'
    os.makedirs(tmp_folder, exist_ok=True)

    # Распаковываем excel как zip в нашу временную папку
    with ZipFile('data/data.xlsx') as excel_container:
        excel_container.extractall(tmp_folder)

    # Переименовываем файл с неверным названием
    wrong_file_path = os.path.join(tmp_folder, 'xl', 'SharedStrings.xml')
    correct_file_path = os.path.join(tmp_folder, 'xl', 'sharedStrings.xml')
    os.rename(wrong_file_path, correct_file_path) 

    # Запаковываем excel обратно в zip и переименовываем в исходный файл
    shutil.make_archive('data/data', 'zip', tmp_folder)
    os.rename('data/data.zip', 'data/data.xlsx')

    book = openpyxl.load_workbook("data/data.xlsx", read_only=True)
    try:
        sheet = book.active
        info = {}
        for item in range(1, sheet.max_row + 1):
            if sheet[item][0].value == number:
                for item2 in range(0, sheet.max_column):
                    if sheet[item][item2].value:
                        info[sheet[3][item2].value] = sheet[item][item2].value
                    else:
                        continue
                break
        if not info:
            return "<b>Hisob topilmadi!</b>"
        else:
            if not db.user_exists(user_id):
                db.add_user(user_id, username, lastname)
            text = "\n".join([f"<b>{key}</b>:  {value}" for key, value in info.items()])
            return text
    finally:
        book.close()


def delete_user(user_id):
    """
        delete user with user_id
    """
    try:
        user_id = int(user_id)
        db.delete_user(user_id)
        return f"User - {user_id} - deleted"
    except Exception as ex:
        return f"Возникла проблема при удалении пользователя: {ex}"


def save_file(filepath):
    """
        accepts the file,
        deletes the old one and saves the new one
    """
    if os.path.exists("data/data.xlsx"):
        try:
            os.remove("data/data.xlsx")
        except Exception as ex:
            return f"<b>error: {str(ex)}</b>"
    with open(f'data/data.xlsx', 'wb') as new_file:
        new_file.write(filepath.read())
        return "<b>Файл сохранен успешно!</b>"
