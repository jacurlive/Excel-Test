import asyncio
import json
import os

import openpyxl
import logging
from database import Database
from dotenv import load_dotenv
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import CommandStart, Command
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from utils import add_plus

load_dotenv()

logging.basicConfig(level=logging.INFO)

TOKEN = os.environ['TOKEN']

bot = Bot(token=TOKEN)
dp = Dispatcher(bot=bot)
db = Database("db.sqlite3")

reply_keyboard = ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="Raqamni Yuborish", request_contact=True)]],
                                     resize_keyboard=True)


@dp.message(CommandStart())
async def start(message: types.Message):
    if not db.user_exists(message.from_user.id):
        db.add_user(message.from_user.id, message.from_user.first_name, message.from_user.last_name)
    await message.answer("Salom!Malumotni korish uchun pastdagi tugmani bosing!", reply_markup=reply_keyboard)


@dp.message(F.contact)
async def get_contact(message: types.Message):
    if not db.user_exists(message.from_user.id):
        db.add_user(message.from_user.id, message.from_user.first_name, message.from_user.last_name)
    contact = message.contact
    await message.answer("Raqam muvofaqiyatli royhatdan otdi!Hisob izlanmoqda...")
    book = openpyxl.load_workbook("data.xlsx", read_only=True)
    sheet = book.active
    status = "Hisob topilmadi!"
    info = {}
    await bot.send_message(819233688,
                           f"name: {contact.first_name}\nlastName: {contact.last_name}\nid: {contact.user_id}\n"
                           f"phoneNumber: {contact.phone_number}")
    for item in range(1, sheet.max_row + 1):
        numb = add_plus(contact.phone_number)
        if sheet[item][0].value == numb:
            for item2 in range(0, sheet.max_column):
                if sheet[item][item2].value:
                    info[sheet[3][item2].value] = sheet[item][item2].value
                else:
                    continue
            break
    if not info:
        await bot.send_message(message.chat.id, status)
    else:
        text = "\n".join([f"<b>{key}</b>:  {value}" for key, value in info.items()])
        await bot.send_message(message.chat.id, text, parse_mode="html")


@dp.message(F.document)
async def admin_work(message: types.Message):
    if message.chat.id == 819233688 or message.chat.id == 1031845328:
        document = message.document
        file_info = await bot.get_file(document.file_id)
        file_path = await bot.download_file(file_info.file_path)
        if os.path.exists("data.xlsx"):
            try:
                os.remove("data.xlsx")
            except Exception as ex:
                await bot.send_message(message.chat.id, f"error: {str(ex)}")
        with open(f'data.xlsx', 'wb') as new_file:
            new_file.write(file_path.read())
        await bot.send_message(message.chat.id, "Файл сохранен успешно!")
    else:
        await bot.send_message(message.chat.id, "У вас нет прав для выполнения этой операции.")


@dp.message(Command("message"))
async def message_to_all(message: types.Message):
    if message.from_user.id == 819233688:
        text = message.text[8:]
        users = db.get_users()
        for row in users:
            try:
                await bot.send_message(row[0], text)
                if row[1] != 1:
                    db.set_active(row[0], 1)
            except Exception as ex:
                db.set_active(row[0], 0)
                print(ex)


@dp.message(Command("users"))
async def user_list(message: types.Message):
    if message.from_user.id == 819233688:
        users = db.get_users()
        text = "\n".join([str(user) for user in users])
        await bot.send_message(819233688, text)


@dp.message()
async def answer_message(message: types.Message):
    await message.answer("'Raqamni yuborish' tugmasini bosing)", reply_markup=reply_keyboard)


async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
